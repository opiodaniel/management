from .basic_ml_objects import BaseDataProcessing, BasePotentialAlgo
from django.apps import apps
from openpyxl import Workbook, load_workbook
import pandas as pd
import string
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.models import User, Group
from .models import Employees, Client, Payment, Land, ClientLand
import random
import string


class AvicAlgo(object):
    def __init__(self, dic):  # to_data_path, target_field
        # print("90004-000 AvibAlgo", dic, '\n', '-'*50)
        try:
            super(AvicAlgo, self).__init__()
        except Exception as ex:
            print("Error 90004-010 AvibDataProcessing:\n"+str(ex), "\n", '-'*50)
        # print("90004-020 AvibAlgo", dic, '\n', '-'*50)


class AvicDataProcessing(BaseDataProcessing, BasePotentialAlgo, AvicAlgo):
    def __init__(self, dic):
        super().__init__(dic)

    def generate_random_plot_number_approved_payment(self, length):
        """Generate a random plot number with the specified length."""
        characters = string.digits
        return ''.join(random.choice(characters) for _ in range(length))

    def generate_random_plot_number_unapproved_payment(self, length):
        """Generate a random plot number with the specified length."""
        characters = string.digits
        return 'Un'.join(random.choice(characters) for _ in range(length))

    def data_upload(self, dic):
        print("90121-1: \n", "=" * 50, "\n", dic, "\n", "=" * 50)

        app_ = 'realestates'
        file_path = self.upload_file(dic)["file_path"]

        print(file_path)

        wb = load_workbook(filename=file_path, read_only=False)
        sheet_names = wb.sheetnames

        try:
            df = pd.read_excel(file_path, sheet_name="Data", header=0)
            print(df.columns)
            print(df)

            df['land_price'] = df['land_price'].fillna(0)
            df['amount_paid'] = df['amount_paid'].fillna(0)
            df['client_contact2'] = df['client_contact2'].fillna('')
            df['employee_contact'] = df['employee_contact'].fillna('')
            df['payment_approved'] = df['payment_approved'].fillna(False)
            df['payment_date'] = pd.to_datetime(df['payment_date'], errors='coerce').fillna(pd.Timestamp('2023-01-01'))

            # Dictionary to track phone numbers and their associated clients
            phone_number_tracker = {}

            for _, row in df.iterrows():
                username_ = str(row['user_name'])

                # Ensure user exists and is associated with an employee
                user, created = User.objects.get_or_create(username=username_, defaults={
                    'email': row['user_email'],
                })
                if created:
                    full_name = string.capwords(str(row["employee_name"]))
                    nnf = full_name.find(" ")
                    first_name = full_name[:nnf]
                    last_name = full_name[nnf + 1:]
                    user.first_name = first_name
                    user.last_name = last_name
                    user.set_password('sql1pass')
                    user.save()

                employee, created = Employees.objects.get_or_create(user=user, defaults={
                    'phone': row['employee_contact']
                })
                if not created:
                    employee.phone = row['employee_contact']
                    employee.save()

                # Track phone numbers and print duplicates
                phone_number = row['client_contact1']
                if phone_number in phone_number_tracker:
                    phone_number_tracker[phone_number].append(row['client_name'])
                else:
                    phone_number_tracker[phone_number] = [row['client_name']]

                # Ensure client exists and is associated with the employee
                client, created = Client.objects.get_or_create(
                    phoneNumber1=phone_number,
                    defaults={
                        'name': row['client_name'],
                        'phoneNumber2': row['client_contact2'],
                        'location': row['client_location'],
                        'employee': employee
                    }
                )
                if not created:
                    client.name = row['client_name']
                    client.phoneNumber2 = row['client_contact2']
                    client.location = row['client_location']
                    client.employee = employee
                    client.save()

                # Only process clients with payments greater than 0 and approved payments
                if row['amount_paid'] > 0 and row['payment_approved']:
                    plot_number = self.generate_random_plot_number_approved_payment(8)

                    land, created = Land.objects.get_or_create(
                        plot_number=plot_number
                    )

                    land.location = row['land_location']
                    land.price = row['land_price']
                    land.available = False
                    land.save()

                    client_land, created = ClientLand.objects.get_or_create(
                        client=client,
                        land=land,
                        defaults={'remaining_amount': land.price}
                    )

                    payment_date = pd.to_datetime(row['payment_date']).date()
                    latest_payment = Payment.objects.filter(client_land=client_land).order_by(
                        'installment_number').last()
                    installment_number = latest_payment.installment_number + 1 if latest_payment else 1

                    payment, created = Payment.objects.get_or_create(
                        client_land=client_land,
                        timestamp=payment_date,
                        defaults={'amount_paid': row['amount_paid'], 'installment_number': installment_number,
                                  'employee': employee, 'approved': row['payment_approved']}
                    )

                    payment.remaining_amount = land.price - row['amount_paid']
                    payment.total_amount = land.price

                    client_land.total_installments += 1
                    client_land.total_amount_paid = row['amount_paid']
                    client_land.remaining_amount = land.price - row['amount_paid']
                    client_land.save()

                    payment.save()

            # Print duplicate phone numbers and associated clients
            for phone_number, clients in phone_number_tracker.items():
                if len(clients) > 1:
                    print(f"Duplicate Phone Number: {phone_number}")
                    print("Clients:")
                    for client in clients:
                        print(f"- {client}")

            return {"status": "ok"}

        except Exception as ex:
            print(f"Error processing sheets: {ex}")
            result = {"status": "error", "message": str(ex)}

        finally:
            wb.close()

        return result



    # def data_upload(self, dic):
    #     print("90121-1: \n", "=" * 50, "\n", dic, "\n", "=" * 50)
    #
    #     app_ = 'realestates'
    #     file_path = self.upload_file(dic)["file_path"]
    #
    #     print(file_path)
    #
    #     wb = load_workbook(filename=file_path, read_only=False)
    #     sheet_names = wb.sheetnames
    #
    #     try:
    #         df = pd.read_excel(file_path, sheet_name="Data", header=0)
    #         print(df.columns)
    #         print(df)
    #
    #         df['land_price'] = df['land_price'].fillna(0)
    #         df['amount_paid'] = df['amount_paid'].fillna(0)
    #         df['client_contact2'] = df['client_contact2'].fillna('')
    #         df['employee_contact'] = df['employee_contact'].fillna('')
    #         df['payment_approved'] = df['payment_approved'].fillna(False)
    #         df['payment_date'] = pd.to_datetime(df['payment_date'], errors='coerce').fillna(pd.Timestamp('2023-01-01'))
    #
    #         # Dictionary to track phone numbers and their associated clients
    #         phone_number_tracker = {}
    #
    #         for _, row in df.iterrows():
    #             username_ = str(row['user_name'])
    #
    #             # Ensure user exists and is associated with an employee
    #             user, created = User.objects.get_or_create(username=username_, defaults={
    #                 'email': row['user_email'],
    #             })
    #             if created:
    #                 full_name = string.capwords(str(row["employee_name"]))
    #                 nnf = full_name.find(" ")
    #                 first_name = full_name[:nnf]
    #                 last_name = full_name[nnf + 1:]
    #                 user.first_name = first_name
    #                 user.last_name = last_name
    #                 user.set_password('sql1pass')
    #                 user.save()
    #
    #             employee, created = Employees.objects.get_or_create(user=user, defaults={
    #                 'phone': row['employee_contact']
    #             })
    #             if not created:
    #                 employee.phone = row['employee_contact']
    #                 employee.save()
    #
    #             # Track phone numbers and print duplicates
    #             phone_number = row['client_contact1']
    #             if phone_number in phone_number_tracker:
    #                 phone_number_tracker[phone_number].append(row['client_name'])
    #             else:
    #                 phone_number_tracker[phone_number] = [row['client_name']]
    #
    #             # Ensure client exists and is associated with the employee
    #             client, created = Client.objects.get_or_create(
    #                 phoneNumber1=phone_number,
    #                 defaults={
    #                     'name': row['client_name'],
    #                     'phoneNumber2': row['client_contact2'],
    #                     'location': row['client_location'],
    #                     'employee': employee
    #                 }
    #             )
    #             if not created:
    #                 client.name = row['client_name']
    #                 client.phoneNumber2 = row['client_contact2']
    #                 client.location = row['client_location']
    #                 client.employee = employee
    #                 client.save()
    #
    #             # Check if payment is approved before creating or getting the Land object
    #             if row['payment_approved']:
    #                 plot_number = self.generate_random_plot_number_approved_payment(8)
    #             else:
    #                 plot_number = self.generate_random_plot_number_unapproved_payment(3)
    #
    #             land, created = Land.objects.get_or_create(
    #                 plot_number=plot_number
    #             )
    #
    #             land.location = row['land_location']
    #             land.price = row['land_price']
    #             land.save()
    #
    #             client_land, created = ClientLand.objects.get_or_create(
    #                 client=client,
    #                 land=land,
    #                 defaults={'remaining_amount': land.price}
    #             )
    #
    #             payment_date = pd.to_datetime(row['payment_date']).date()
    #             latest_payment = Payment.objects.filter(client_land=client_land).order_by('installment_number').last()
    #             installment_number = latest_payment.installment_number + 1 if latest_payment else 1
    #
    #             payment, created = Payment.objects.get_or_create(
    #                 client_land=client_land,
    #                 timestamp=payment_date,
    #                 defaults={'amount_paid': row['amount_paid'], 'installment_number': installment_number,
    #                           'employee': employee, 'approved': row['payment_approved']}
    #             )
    #
    #             client_land.total_installments += 1
    #             client_land.total_amount_paid += row['amount_paid']
    #             client_land.remaining_amount -= row['amount_paid']
    #             client_land.save()
    #
    #         # Calculate monthly totals
    #         df['month'] = df['payment_date'].dt.to_period('M')
    #         monthly_totals = df.groupby('month')['amount_paid'].sum().reset_index()
    #         monthly_totals['date'] = monthly_totals['month'].dt.to_timestamp()
    #
    #         # Update or create entries in TotalAmount model
    #         for _, row in monthly_totals.iterrows():
    #             date = row['date'].date()
    #             amount = row['amount_paid']
    #             TotalAmount.objects.update_or_create(date=date, defaults={'amount': amount})
    #
    #         # Print duplicate phone numbers and associated clients
    #         for phone_number, clients in phone_number_tracker.items():
    #             if len(clients) > 1:
    #                 print(f"Duplicate Phone Number: {phone_number}")
    #                 print("Clients:")
    #                 for client in clients:
    #                     print(f"- {client}")
    #
    #         return {"status": "ok"}
    #
    #     except Exception as ex:
    #         print(f"Error processing sheets: {ex}")
    #         result = {"status": "error", "message": str(ex)}
    #
    #     finally:
    #         wb.close()
    #
    #     return result
    # current
    # def data_upload(self, dic):
    #     print("90121-1: \n", "=" * 50, "\n", dic, "\n", "=" * 50)
    #
    #     app_ = 'realestates'
    #     file_path = self.upload_file(dic)["file_path"]
    #
    #     print(file_path)
    #
    #     wb = load_workbook(filename=file_path, read_only=False)
    #     sheet_names = wb.sheetnames
    #
    #     try:
    #         df = pd.read_excel(file_path, sheet_name="Data", header=0)
    #         print(df.columns)
    #         print(df)
    #
    #         df['land_price'] = df['land_price'].fillna(0)
    #         df['amount_paid'] = df['amount_paid'].fillna(0)
    #         df['client_contact2'] = df['client_contact2'].fillna('')
    #         df['employee_contact'] = df['employee_contact'].fillna('')
    #         df['payment_approved'] = df['payment_approved'].fillna(False)
    #         df['payment_date'] = pd.to_datetime(df['payment_date'], errors='coerce').fillna(pd.Timestamp('2023-01-01'))
    #
    #         # Dictionary to track phone numbers and their associated clients
    #         phone_number_tracker = {}
    #
    #         for _, row in df.iterrows():
    #             username_ = str(row['user_name'])
    #
    #             # Ensure user exists and is associated with an employee
    #             user, created = User.objects.get_or_create(username=username_, defaults={
    #                 'email': row['user_email'],
    #             })
    #             if created:
    #                 full_name = string.capwords(str(row["employee_name"]))
    #                 nnf = full_name.find(" ")
    #                 first_name = full_name[:nnf]
    #                 last_name = full_name[nnf + 1:]
    #                 user.first_name = first_name
    #                 user.last_name = last_name
    #                 user.set_password('sql1pass')
    #                 user.save()
    #
    #             employee, created = Employees.objects.get_or_create(user=user, defaults={
    #                 'phone': row['employee_contact']
    #             })
    #             if not created:
    #                 employee.phone = row['employee_contact']
    #                 employee.save()
    #
    #             # Track phone numbers and print duplicates
    #             phone_number = row['client_contact1']
    #             if phone_number in phone_number_tracker:
    #                 phone_number_tracker[phone_number].append(row['client_name'])
    #             else:
    #                 phone_number_tracker[phone_number] = [row['client_name']]
    #
    #             # Ensure client exists and is associated with the employee
    #             client, created = Client.objects.get_or_create(
    #                 phoneNumber1=phone_number,
    #                 defaults={
    #                     'name': row['client_name'],
    #                     'phoneNumber2': row['client_contact2'],
    #                     'location': row['client_location'],
    #                     'employee': employee
    #                 }
    #             )
    #             if not created:
    #                 client.name = row['client_name']
    #                 client.phoneNumber2 = row['client_contact2']
    #                 client.location = row['client_location']
    #                 client.employee = employee
    #                 client.save()
    #
    #             # Check if payment is approved before creating or getting the Land object
    #             if row['payment_approved']:
    #                 plot_number = self.generate_random_plot_number_approved_payment(8)
    #             else:
    #                 plot_number = self.generate_random_plot_number_unapproved_payment(3)
    #
    #             land, created = Land.objects.get_or_create(
    #                 plot_number=plot_number
    #             )
    #
    #             land.location = row['land_location']
    #             land.price = row['land_price']
    #             land.save()
    #
    #             client_land, created = ClientLand.objects.get_or_create(
    #                 client=client,
    #                 land=land,
    #                 defaults={'remaining_amount': land.price}
    #             )
    #
    #             payment_date = pd.to_datetime(row['payment_date']).date()
    #             latest_payment = Payment.objects.filter(client_land=client_land).order_by('installment_number').last()
    #             installment_number = latest_payment.installment_number + 1 if latest_payment else 1
    #
    #             payment, created = Payment.objects.get_or_create(
    #                 client_land=client_land,
    #                 timestamp=payment_date,
    #                 defaults={'amount_paid': row['amount_paid'], 'installment_number': installment_number,
    #                           'employee': employee, 'approved': row['payment_approved']}
    #             )
    #
    #             client_land.total_installments += 1
    #             client_land.total_amount_paid += row['amount_paid']
    #             client_land.remaining_amount -= row['amount_paid']
    #             client_land.save()
    #
    #         # Print duplicate phone numbers and associated clients
    #         for phone_number, clients in phone_number_tracker.items():
    #             if len(clients) > 1:
    #                 print(f"Duplicate Phone Number: {phone_number}")
    #                 print("Clients:")
    #                 for client in clients:
    #                     print(f"- {client}")
    #
    #         return {"status": "ok"}
    #
    #     except Exception as ex:
    #         print(f"Error processing sheets: {ex}")
    #         result = {"status": "error", "message": str(ex)}
    #
    #     finally:
    #         wb.close()
    #
    #     return result



# Duplicate Phone Number: 0742719473
# Clients:
# - Patricia
# - Patricia
# Duplicate Phone Number: 0752520074
# Clients:
# - Konsolet
# - Konsolet
# Duplicate Phone Number: 0758859629
# Clients:
# - Sumuel
# - Sumuel
# Duplicate Phone Number: 07599644729
# Clients:
# - Kabuye ssekito
# - Kabuye ssekito






