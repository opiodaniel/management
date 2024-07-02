from django.shortcuts import render, redirect,  get_object_or_404

from django.contrib import messages
from django.contrib.auth.models import User,auth

from .models import Client, Employees, Payment,  Company, Commission, Land, ClientLand
from .forms import (ClientForm, ClientEditForm, EmployeeLoginForm, RegistrationForm,
                    ProfileEditForm, UserEditForm, PaymentForm, LandForm, ClientLandForm)

from django.contrib.auth import authenticate, login

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse

from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash

from django.core.mail import send_mail

from datetime import date
from django.utils import timezone

from django.views.generic.edit import CreateView, UpdateView, DeleteView
from .tasks import remove_inactive_clients


from django.views.generic.edit import FormView

from django.urls import reverse, reverse_lazy

from django.db.models import Sum
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.contrib.auth.models import User
import json
from datetime import datetime, timedelta, time
from decimal import Decimal
from django.http import HttpResponse
from .models import EmployeePaymentRecord
from django.db import transaction
from django.contrib.auth import logout
from django.views.generic import View
from django import forms

from django.apps import apps
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Q, Prefetch
import logging
import pandas as pd
from django.http import HttpResponseForbidden


from .apps_general_functions import activate_obj_function
from django.db.models.fields.related import ForeignKey
import numbers
from django.apps import apps
from django.conf import settings
from django.core.files.storage import FileSystemStorage
import os
import calendar
from django.core.cache import cache
from django.db.models import Sum, F, ExpressionWrapper, IntegerField
from django.utils.timezone import now
import openpyxl
from openpyxl.utils import get_column_letter
from django.views.decorators.http import require_POST


# ======== LOGIN PAGE ========
def login_page(request):
    company = Company.objects.get(id=1)
    company_logo_url = company.company_logo.url
    arg = {
           'company_logo_url': company_logo_url
           }
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Authenticate user
        user = authenticate(request, username=username, password=password)
        # user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            if user.is_staff:  # Check if user is admin
                # return redirect('admin_dashboard')  # Redirect admin to admin dashboard
                # print(user)
                # print(user.id)
                # messages.info(request, f"You are now logged in as {username}.")
                return redirect(reverse('realestates:admin_dashboard', args=[user.id]))
            else:
                # messages.info(request, f"You are now logged in as {username}.")
                return redirect(reverse('realestates:employee_dashboard'))
        else:
            messages.error(request, "Invalid username or password.")

        # if user is not None:
        #     # Login user
        #     login(request, user)
        #     messages.success(request, 'You have successfully logged in.')
        #     return redirect('home')  # Redirect to home page after successful login
        # else:
        #     messages.error(request, 'Invalid username or password.')
        #     return redirect('login')  # Redirect back to login page with error message

    # If request method is GET, render the login page
    return render(request, 'realestates/registration/login.html', arg)


# ======= CALCULATES MONTHLY SALE BASING ON THE PAYMENT MODEL =========
def get_total_sales_for_month(request, year, month):
    # Define the start date of the month
    start_date = datetime(year, month, 1)
    # Define the end date of the month
    end_date = datetime(year, month, 1) + timedelta(days=32)
    end_date = datetime(end_date.year, end_date.month, 1) - timedelta(days=1)
    # Calculate the total sales for the month
    total_sales = Payment.objects.filter(timestamp__range=[start_date, end_date]).aggregate(total_sales=Sum('amount_paid'))['total_sales']
    # print(total_sales)
    # If no sales, return 0
    if total_sales is None:
        total_sales = 0
    return total_sales


def get_total_sales_for_previous_months(request):
    total_sales_previous_months = []
    total_sales_previous_months_sum = 0
    # Get the current year and month
    current_year = datetime.now().year
    current_month = datetime.now().month
    # Iterate over previous months
    for month in range(1, current_month):
        total_sales = get_total_sales_for_month(request, current_year, month)
        total_sales_previous_months.append((datetime(current_year, month, 1), total_sales))
        total_sales_previous_months_sum += total_sales
    # Calculate total sales for the current month
    total_sales_current_month = get_total_sales_for_month(request, current_year, current_month)
    # print('=== total_sales_current_month === ', total_sales_current_month)
    total_sales_previous_months.append((datetime(current_year, current_month, 1), total_sales_current_month))
    total_sales_previous_months_sum += total_sales_current_month
    # print('=== total_sales_previous_months_sum === ', total_sales_previous_months_sum)
    return total_sales_previous_months, total_sales_previous_months_sum


# ========== ADMIN DASHBOARD ============
@login_required
def admin_dashboard(request, admin_id):

    default_threshold_input = 10  # employees with more than 5 clients who have made some payment will be displayed

    if request.method == 'POST' and 'confirm_client_payment' in request.POST:
        client_id = request.POST.get('client_id')
        client = Client.objects.get(id=client_id)
        return redirect(reverse('realestates:employee_dashboard', args=[request.user.id]))

    authenticated_admin = request.user.id

    if authenticated_admin != int(admin_id):
        return redirect(reverse('realestates:admin_dashboard', args=[request.user.id]))

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    current_month = datetime.now().month
    start_date = datetime(datetime.now().year, current_month, 1)
    end_date = datetime(datetime.now().year, current_month + 1, 1) - timedelta(days=1)

    total_amount_month = Payment.objects.filter(timestamp__range=[start_date, end_date]).aggregate(
        total_amount_month=Sum('amount_paid'))['total_amount_month']

    if total_amount_month is None:
        total_amount_month = 0
    total_amount_month_ = total_amount_month
    total_amount_month = '{:,}'.format(total_amount_month_)

    total_sales_previous_months, total_sales_previous_months_sum = get_total_sales_for_previous_months(request)

    monthly_data = []
    month_name_to_number = {
        "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
        "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12
    }
    for month, total_sales in total_sales_previous_months:
        month_number = month_name_to_number[month.strftime('%B')]
        selected_year = month.strftime('%Y')
        amount_details_in_month = Payment.objects.filter(timestamp__month=month_number, timestamp__year=selected_year).filter(approved=True)
        monthly_data.append({
            'date': month.strftime('%B %Y'),
            'amount': "{:,}".format(total_sales),
            'month_detail': amount_details_in_month,
        })
    total_sales_previous_months_sum_ = total_sales_previous_months_sum
    total_sales_previous_months_sum = '{:,}'.format(total_sales_previous_months_sum_)

    today = date.today()

    total_amount_made_today = Payment.objects.filter(timestamp=today).aggregate(total_amount_made_today=Sum('amount_paid'))['total_amount_made_today']

    if total_amount_made_today is None:
        total_amount_made_today = 0

    total_amount_previous_days = Payment.objects.filter(timestamp__lt=today).aggregate(total_amount_previous_days=Sum('amount_paid'))['total_amount_previous_days']

    if total_amount_previous_days is None:
        total_amount_previous_days = 0
    total_sale_for_today_previous_days = total_amount_made_today + total_amount_previous_days

    employees = Employees.objects.filter(is_administrator=False)

    payments = Payment.objects.filter(approved=False).order_by('approved', '-client_land__purchase_date')
    total_num_clients_ = Client.objects.all().count()
    total_number_employees = Employees.objects.filter(is_administrator=False).count()

    # Calculate data for the pie chart
    clients_with_land_and_payment = ClientLand.objects.filter(payments__isnull=False).distinct().count()
    clients_without_land = Client.objects.filter(client_lands__isnull=True).count()

    context = {
        'admin_id': admin_id,
        'employees': employees,
        'payments': payments,
        'total_number_employees': total_number_employees,
        'total_num_clients_': total_num_clients_,
        'total_amount_made_today': total_amount_made_today,
        'total_sale_for_today_previous_days': total_sale_for_today_previous_days,
        'company': company,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
        'total_amount_month': total_amount_month,
        'total_sales_previous_months_sum': total_sales_previous_months_sum,
        'monthly_data': monthly_data,
        "default_threshold_input": default_threshold_input,
        'clients_with_land_and_payment': clients_with_land_and_payment,
        'clients_without_land': clients_without_land,
    }
    return render(request, 'realestates/admin/admin_dashboard.html', context)


# ========= LISTS OF FREE CLIENTS =========
@login_required
def free_clients(request):
    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    # Calculate the last Sunday at midnight
    now = timezone.now()
    last_sunday = now - timedelta(days=(now.weekday() + 1) % 7)
    last_sunday_midnight = datetime.datetime.combine(last_sunday, datetime.time.min)

    # Print the calculated expiry date for debugging purposes
    # print(last_sunday_midnight)

    # Retrieve expired clients
    query = request.GET.get('q')
    if query:
        all_expired_clients = Client.objects.filter(
            Q(name__icontains=query) | Q(phoneNumber1__icontains=query),
            date__lt=last_sunday_midnight,
            client_lands__isnull=True
        ).order_by('date')  # Order by 'date' or any other field
    else:
        all_expired_clients = Client.objects.filter(
            date__lt=last_sunday_midnight
        ).exclude(client_lands__isnull=False).order_by('date')

    # Set up pagination
    paginator = Paginator(all_expired_clients, 10)  # Show 10 clients per page
    page = request.GET.get('page')

    try:
        clients = paginator.page(page)
    except PageNotAnInteger:
        clients = paginator.page(1)
    except EmptyPage:
        clients = paginator.page(paginator.num_pages)

    employees = Employees.objects.all()
    context = {
        'clients': clients,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
        'admin_id': admin_id,
        'query': query,
        'employees': employees,
    }
    return render(request, 'realestates/admin/free_clients.html', context)


@login_required
def free_clients(request):
    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    # Calculate the last Sunday at midnight
    now = timezone.now()
    last_sunday = now - timedelta(days=(now.weekday() + 1) % 7)
    last_sunday_midnight = datetime.combine(last_sunday, time.min)

    # Print the calculated expiry date for debugging purposes
    print(last_sunday_midnight)

    # Retrieve expired clients
    query = request.GET.get('q')
    if query:
        all_expired_clients = Client.objects.filter(
            Q(name__icontains=query) | Q(phoneNumber1__icontains=query),
            date__date__lt=last_sunday_midnight,
            client_lands__isnull=True
        ).order_by('date')  # Order by 'date' or any other field
    else:
        all_expired_clients = Client.objects.filter(
            date__date__lt=last_sunday_midnight
        ).exclude(client_lands__isnull=False).order_by('date')

    # Set up pagination
    paginator = Paginator(all_expired_clients, 10)  # Show 10 clients per page
    page = request.GET.get('page')

    try:
        clients = paginator.page(page)
    except PageNotAnInteger:
        clients = paginator.page(1)
    except EmptyPage:
        clients = paginator.page(paginator.num_pages)

    employees = Employees.objects.all()
    context = {
        'clients': clients,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
        'admin_id': admin_id,
        'query': query,
        'employees': employees,
    }
    return render(request, 'realestates/admin/free_clients.html', context)


# ======== DOWNLOAD EXCEL FILE CONTAINING LISTS OF FREE CLIENTS =========
@login_required
def download_free_clients(request):
    query = request.GET.get('q')
    expiry_date = timezone.now().date() - timedelta(days=0)

    if query:
        expired_clients = Client.objects.filter(
            Q(name__icontains=query) | Q(phoneNumber1__icontains=query),
            date__date__lt=expiry_date,
            client_lands__isnull=True
        ).order_by('name')
    else:
        expired_clients = Client.objects.filter(
            date__date__lt=expiry_date,
            client_lands__isnull=True
        ).order_by('name')

    # Create an in-memory workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Free Clients'

    # Write the header row
    headers = ['Name', 'Contact']
    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    for row_num, client in enumerate(expired_clients, 2):
        worksheet.cell(row=row_num, column=1, value=client.name)
        worksheet.cell(row=row_num, column=2, value=client.phoneNumber1)

    # Adjust column widths
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 20

    # Save the workbook to an in-memory file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=free_clients.xlsx'
    workbook.save(response)

    return response


# ========== ADMIN TO ASSIGN FREE CLIENTS TO EMPLOYEES ==========
@login_required
def assign_client(request, client_id):
    if request.method == 'POST':
        client = get_object_or_404(Client, id=client_id)
        employee_id = request.POST.get('employee_id')
        employee = get_object_or_404(Employees, id=employee_id)
        client.employee = employee
        client.date = timezone.now()
        client.save()
        return redirect('realestates:free_clients')


# ======= EMPLOYEES TO CLAIM FREE CLIENTS.(7 WEEK EXPIRED CLIENTS WITHOUT ANY PAYMENT ) ===========
@login_required
def claim_free_client(request):
    if request.method == "POST":
        phone_number = request.POST.get("phone_number")
        print(phone_number)

        try:
            client = Client.objects.get(phoneNumber1=phone_number)
            expiry_date = timezone.now().date() - timedelta(days=0)

            if client.date.date() < expiry_date and not client.client_lands.exists():
                client.employee = request.user.employee
                client.date = timezone.now()
                client.save()
                response = {
                    'success': True,
                    'message': f'Client {client.name} has been successfully claimed.'
                }
            else:
                response = {
                    'success': False,
                    'message': 'Client is not free or is already claimed.'
                }
        except Client.DoesNotExist:
            response = {
                'success': False,
                'message': 'Client with that phone number does not exist.'
            }

        return JsonResponse(response)


# ========== LEADS TO THE PAYMENT PAGE FOR THE EMPLOYEE
@login_required
def pay_employee(request):
    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    query = request.GET.get('q')
    employees = Employees.objects.filter(is_administrator=False,
                                         clients__client_lands__payments__isnull=False).distinct().order_by('-created')

    if query:
        employees = employees.filter(user__first_name__icontains=query) | employees.filter(user__last_name__icontains=query)

    paginator = Paginator(employees, 50)  # Show 50 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'admin_id': admin_id,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
        'query': query,
    }

    return render(request, 'realestates/admin/pay_employee.html', context)


# ========= EMPLOYEE CLIENTS WHO MADE SOME PAYMENT FOR THE LAND  =========
@login_required
def employee_clients_made_payment(request, employee_id):
    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    employee = get_object_or_404(Employees, id=employee_id)
    client_lands = ClientLand.objects.filter(client__employee=employee).select_related('client', 'land')
    client_commissions = Commission.objects.filter(employee=employee).select_related('client')

    client_data = []
    for client_land in client_lands:
        commission = client_commissions.filter(client=client_land.client).first()
        total_commission = commission.total_commission if commission else 0
        amount_paid = EmployeePaymentRecord.objects.filter(employee=employee, client=client_land.client).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        balance = total_commission - amount_paid

        client_data.append({
            'client_land': client_land,
            'commission': total_commission,
            'amount_paid': amount_paid,
            'balance': balance,
        })

    context = {
        'admin_id': admin_id,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
        'employee_': employee,
        'client_data': client_data,
    }

    return render(request, 'realestates/admin/employee_clients_made_payment.html', context)


# ========= CONFIRM EMPLOYEE PAYMENT =========
@require_POST
@login_required
def approve_employee_payment(request, employee_id, client_id):
    if request.method == "POST":
        amount_paid_str = request.POST.get('amount_paid', '0').replace(',', '')
        amount_paid = int(amount_paid_str)
        employee = get_object_or_404(Employees, id=employee_id)
        client = get_object_or_404(Client, id=client_id)

        try:
            commission = Commission.objects.get(employee=employee, client=client)
            total_commission = commission.total_commission
        except Commission.DoesNotExist:
            total_commission = 0

        amount_already_paid = EmployeePaymentRecord.objects.filter(employee=employee, client=client).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        remaining_commission = total_commission - amount_already_paid

        if amount_paid > remaining_commission:
            messages.error(request, f"Amount to be paid exceeds the remaining commission of {'{:,}'.format(remaining_commission)}.")
        else:
            # Proceed with the payment
            EmployeePaymentRecord.objects.create(
                employee=employee,
                client=client,
                total_commission=total_commission,
                amount_paid=amount_paid,
                balance=remaining_commission - amount_paid
            )
            messages.success(request, f"Successfully paid {'{:,}'.format(amount_paid)} to {employee.user.get_full_name()} for client {client.name}.")

        return redirect('realestates:employee_clients_made_payment', employee_id=employee.id)

    return redirect('realestates:employee_clients_made_payment', employee_id=employee_id)


# ========== CLIENT LIST ============
@login_required
def client_list(request):

    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    expiry_date = timezone.now().now().date() - timedelta(days=7)
    # Filter clients who have not exceeded the expiry date
    # active_clients = Client.objects.filter(date__date__gte=expiry_date).order_by('-date')

    active_clients = Client.objects.all().order_by('-date')

    all_clients = active_clients

    # Handle search query
    query = request.GET.get('q')
    if query:
        all_clients = all_clients.filter(phoneNumber1__icontains=query)  # Adjust field ('name') based on search criteria
    # Apply pagination
    paginator = Paginator(all_clients, 50)  # Display 50 clients per page
    page_number = request.GET.get('page')
    try:
        clients = paginator.page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        clients = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results
        clients = paginator.page(paginator.num_pages)

    print(clients)

    context = {
        'clients': clients,
        'admin_id': admin_id,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
    }

    return render(request, 'realestates/admin/client_list.html', context)


# ============= EDIT USER PROFILE =================
def edit_user_profile_new(request):
    # company_obj = WebSiteCompany(request, web_company_id=7).site_company()
    profile = Employees.objects.filter(user=request.user)
    if not profile:
        profile = Employees.objects.create(user=request.user)
    if request.method == 'POST':
        user_form = UserEditForm(instance=request.user, data=request.POST)
        profile_form = ProfileEditForm(instance=request.user.employee, data=request.POST, files=request.FILES)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            if request.user.is_staff:  # Check if user is admin
                # return redirect('admin_dashboard')  # Redirect admin to admin dashboard
                # print(user)
                # print(user.id)
                return redirect(reverse('realestates:admin_dashboard', args=[request.user.id]))
            else:
                return redirect(reverse('realestates:employee_dashboard', args=[request.user.id]))
    else:
        user_form = UserEditForm(instance=request.user)
        profile_form = ProfileEditForm(instance=request.user.employee)
    return render(request, 'realestates/registration/edit_user_profile.html', {'user_form': user_form,
                                                                                        'profile_form': profile_form, })


# ============= CHANGE USER PASSWORD =================
@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(data=request.POST, user=request.user)
        for field in form:
            print("Field Error:", field.name, field.errors)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect(reverse('realestates:login_page'))
        else:
            print("Field Error:", field.name, field.errors)
            messages.error(request, field.errors)
            return redirect(reverse('realestates:change_password'))
    else:
        form = PasswordChangeForm(user=request.user)

        args = {'form': form}
        return render(request, 'realestates/registration/change_password.html', args)


# ============ RECORDING CLIENT PAYMENT ==================
def record_payment(request, client_id):

    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    client = get_object_or_404(Client, id=client_id)
    # client_land = client.client_lands.first()  # Get the first associated land, if any
    # Get the first associated land where the payment is not complete
    client_land = client.client_lands.filter(payment_complete=False).first()

    # Ensure transactions is always a queryset
    if client_land:
        transactions = Payment.objects.filter(client_land=client_land)
    else:
        transactions = Payment.objects.none()

    # Calculate the total amount paid
    total_amount_paid = transactions.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0

    # Calculate the balance only if client_land is not None
    if client_land:
        land_price = client_land.land.price
        balance = land_price - total_amount_paid
    else:
        balance = None

    if request.method == 'POST':
        if not client_land or client_land.payment_complete:
            # Handle ClientLandForm submission
            landform = ClientLandForm(request.POST)
            if landform.is_valid():
                client_land = landform.save()
                return redirect('realestates:record_payment', client_id=client.id)
        else:
            # Handle PaymentForm submission
            form = PaymentForm(request.POST)
            if form.is_valid():
                payment = form.save(commit=False)
                payment.client_land = client_land
                payment.approved_by = request.user.employee if request.user.employee.is_administrator else None
                payment.save()
                return redirect('realestates:record_payment', client_id=client.id)
    else:
        landform = ClientLandForm(initial={'client': client})

        initial_data = {'employee': request.user.employee if request.user.employee.is_administrator else None}
        if client_land:
            initial_data['client_land'] = client_land
        form = PaymentForm(initial=initial_data)

    available_lands = Land.objects.filter(available=True)
    addlandform = LandForm()

    context = {
        'form': form,
        'landform': landform,
        'addlandform': addlandform,
        'client': client,
        'client_land': client_land,
        'transactions': transactions,
        'total_amount_paid': total_amount_paid,
        'balance': balance,
        'available_lands': available_lands,
        'admin_id': admin_id,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
    }
    return render(request, 'realestates/admin/record_payment.html', context)


# ============ EDIT PAYMENT MADE BY CLIENT =================
def edit_payment(request, payment_id):
    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    payment = get_object_or_404(Payment, id=payment_id)
    client = payment.client_land.client

    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        # print(form.errors)
        if form.is_valid():
            form.save()
            return redirect('realestates:record_payment', client_id=client.id)
    else:
        form = PaymentForm(instance=payment)

    context = {
        'form': form,
        'payment': payment,
        'admin_id': admin_id,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
    }
    return render(request, 'realestates/admin/edit_payment.html', context)


# ============== ADD NEW LAND INTO THE SYSTEM ================
def add_land(request):
    if request.method == 'POST':
        form = LandForm(request.POST)
        if form.is_valid():
            land = form.save()
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'errors': form.errors})
    return JsonResponse({'status': 'invalid method'}, status=405)


# ============= CLIENTS WITH LAND(MADE COMPLETE PAYMENT FOR THE LAND) ==============
def clients_with_lands(request):
    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    clients = Client.objects.filter(client_lands__payment_complete=True).distinct().order_by('-date')

    # Handle search query
    search_query = request.GET.get('search', '')
    if search_query:
        clients = clients.filter(phoneNumber1__icontains=search_query)

    paginator = Paginator(clients, 10)  # Show 10 clients per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'admin_id': admin_id,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
    }

    return render(request, 'realestates/admin/clients_with_lands.html', context)


# ============== TRANSACTION HISTORY FOR THE CLIENTS WHO ALREADY BOUGHT LAND ===============
def land_transaction_history(request, land_id):
    admin_id = request.user.id

    profile_pic_url = ""
    try:
        admin = Employees.objects.get(id=request.user.id)
        if admin.profile_pic:
            profile_pic_url = admin.profile_pic.url
    except Employees.DoesNotExist:
        admin = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    client_land = get_object_or_404(ClientLand, id=land_id)
    transactions = Payment.objects.filter(client_land=client_land)
    total_amount_paid = transactions.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    context = {
        'client_land': client_land,
        'transactions': transactions,
        'total_amount_paid': total_amount_paid,
        'admin_id': admin_id,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,

    }
    return render(request, 'realestates/admin/land_transaction_history.html',
                  context)


# ================ EMPLOYEE DASHBOARD ==============
@login_required
def employee_dashboard(request):

    try:
        employee = request.user.employee
    except Employees.DoesNotExist:
        return HttpResponseForbidden("You are not authorized to view this page.")

    # Retrieve the authenticated employee
    authenticated_employee = request.user.id

    profile_pic_url = ""
    try:
        employee = request.user.employee
        if employee.profile_pic:
            profile_pic_url = employee.profile_pic.url
    except Employees.DoesNotExist:
        employee = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    payments = Payment.objects.filter(employee=employee).order_by('approved', '-client_land__client__date')

    total_number_of_clients = employee.total_clients()
    total_approved_clients = employee.total_approved_clients()
    total_appending_clients = employee.total_appending_clients()

    clients = Client.objects.filter(employee=employee).exclude(client_lands__isnull=False).order_by('-date')

    commission_earned_per_client = Commission.objects.filter(employee=employee)
    total_commission = Commission.objects.filter(employee=employee).aggregate(Sum('total_commission'))[
                           'total_commission__sum'] or 0
    total_paid = employee.payment_records.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    balance = total_commission - total_paid

    form = ClientForm()

    context = {
        'employee': employee,
        'profile_pic_url': profile_pic_url,
        'company_logo_url': company_logo_url,
        'clients': clients,
        'total_number_of_clients': total_number_of_clients,
        'total_approved_clients': total_approved_clients,
        'total_appending_clients': total_appending_clients,
        'payments': payments,
        'commission_earned_per_client': commission_earned_per_client,
        'total_commission': total_commission,
        'total_paid': total_paid,
        'balance': balance,
        'form': form,
    }

    # Render the employee dashboard template
    return render(request, 'realestates/employee/employee_dashboard.html', context)


# ========== ENTER NEW CLIENT ==========
@login_required
def add_client(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save(commit=False)
            employee = Employees.objects.get(user=request.user)
            client.employee = employee
            client.date = timezone.now()
            client.save()
            return JsonResponse({'success': True})
        else:
            errors = form.errors.as_json()
            return JsonResponse({'success': False, 'error': errors})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


# ========== EDIT CLIENT ==========
@login_required
def edit_client(request, pk):

    client_info = get_object_or_404(Client, id=pk)

    if request.method == 'POST':
        client_form = ClientEditForm(data=request.POST, instance=client_info)
        if client_form.is_valid():
            client_form.save()
            return redirect(reverse('realestates:employee_dashboard'))
    else:
        # Populate forms with existing data
        client_form = ClientEditForm(instance=client_info)

    return render(request, 'realestates/employee/edit_client.html', {'form': client_form})


# ========== DISPLAYS EMPLOYEES WITH THEIR CLIENTS =============
@login_required
def employees_client(request):
    employees = Employees.objects.filter(is_administrator=False)

    # Prefetch related data for optimized querying
    employees = employees.prefetch_related(
        Prefetch(
            'employee_payments',
            queryset=Payment.objects.select_related('client_land__client')
        ),
        Prefetch(
            'clients',
            queryset=Client.objects.all()
        )
    )

    dic_ = {}

    for employee in employees:
        approved_clients = []
        pending_clients = []
        all_clients = list(employee.clients.all())

        for payment in employee.employee_payments.all():
            client = payment.client_land.client
            if payment.approved:
                approved_clients.append(client)
            else:
                pending_clients.append(client)

        # Remove clients with payments from all_clients to get clients with no payments
        clients_with_no_payments = [client for client in all_clients if client not in approved_clients and client not in pending_clients]

        dic_[employee.user.username] = {
            'approved_clients': [client.name for client in approved_clients],
            'pending_clients': [client.name for client in pending_clients],
            'clients_with_no_payments': [client.name for client in clients_with_no_payments],
            'total_clients': [client.name for client in all_clients]
        }

    return JsonResponse(dic_)


# ========= EMPLOYEE CLIENTS WHO MADE SOME PAYMENT FOR THE LAND  =========
@login_required
def employee_pay_breakdown(request):

    try:
        employee = request.user.employee
    except Employees.DoesNotExist:
        return HttpResponseForbidden("You are not authorized to view this page.")

    employee_id = request.user.employee.id

    profile_pic_url = ""
    try:
        employee = Employees.objects.get(id=request.user.id)
        if employee.profile_pic:
            profile_pic_url = employee.profile_pic.url
    except Employees.DoesNotExist:
        employee = None

    company_logo_url = ""
    try:
        company = Company.objects.get(id=1)
        if company.company_logo:
            company_logo_url = company.company_logo.url
    except Company.DoesNotExist:
        company = None

    employee = get_object_or_404(Employees, id=employee_id)
    client_lands = ClientLand.objects.filter(client__employee=employee).select_related('client', 'land')
    client_commissions = Commission.objects.filter(employee=employee).select_related('client')

    client_data = []
    for client_land in client_lands:
        commission = client_commissions.filter(client=client_land.client).first()
        total_commission = commission.total_commission if commission else 0
        amount_paid = EmployeePaymentRecord.objects.filter(employee=employee, client=client_land.client).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        balance = total_commission - amount_paid

        client_data.append({
            'client_land': client_land,
            'commission': total_commission,
            'amount_paid': amount_paid,
            'balance': balance,
        })

    context = {
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
        'employee': employee,
        'client_data': client_data,
    }

    return render(request, 'realestates/employee/employee_pay_breakdown.html', context)


# =========== SEND EMAIL WHEN NEW EMPLOYEE/DISTRIBUTOR IS CREATED ==========
def email_message(semail, username,  type):
    if type == 'register':
        # print("0044444444444444")
        email_from = 'noreply@drbaranes.com'
        subject = 'Registering in Century Properties & Real Estates Ltd'
        body = 'You were registered. Temporal password: sql1pass and username: '+username +' Please login and Update/Edit your profile .' \
               'https://centuryproperties.pythonanywhere.com/'
        # print("005555555555555")

    send_mail(subject, body, email_from, [semail], fail_silently=False)


# ============ REGISTER MORE EMPLOYEES/DISTRIBUTOR ==================
@login_required
def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        # print(form.errors)
        if form.is_valid():
            cd = form.cleaned_data
            semail = cd['email']
            username = cd['username']
            email_message(semail, username, 'register')
            form.save()
            return redirect(reverse('realestates:admin_dashboard', args=[request.user.id]))

    form = RegistrationForm()
    admin_id = request.user.id
    admin = Employees.objects.get(id=request.user.id)
    profile_pic_url = admin.profile_pic.url
    company = Company.objects.get(id=1)
    company_logo_url = company.company_logo.url
    employees = Employees.objects.all().filter(is_administrator=False)
    context = {
        'form': form,
        'admin_id': admin_id,
        'employees': employees,
        'company_logo_url': company_logo_url,
        'profile_pic_url': profile_pic_url,
    }

    return render(request, 'realestates/registration/reg_form.html', context)


# ========= LOGOUT FUNCTION ===============
class CustomLogoutView(View):
    def post(self, request):
        # Handle post request if needed
        logout(request)
        messages.success(request, 'You have been logged out successfully.')
        return redirect('realestates:login_page')  # Adjust the redirect URL as needed

    def get(self, request):
       pass


# ============= CLEANING THE DATABASE ================
def truncate_model(request):
    # print('9015 params')
    # print(params)
    dic = request.POST['dic']
    print(dic)
    dic_ = eval(dic)
    app_ = dic_['app']
    model_name_ = dic_['model_name']
    try:
        model = apps.get_model(app_label=app_, model_name=model_name_)
        print(model)
        model.truncate()
    except Exception as ex:
        print("9025 " + str(ex))

    result = "Data truncated"
    return result


# ============  DOWNLOAD EXCEL CONTAINING CLIENTS WHO HAVEN'T MADE ANY PAYMENT ===============
def export_unapproved_payments(request):
    one_week_ago = timezone.now() - timezone.timedelta(days=1)
    unapproved_payments = Payment.objects.filter(approved=False, timestamp__lte=one_week_ago)
    # print(unapproved_payments)

    data = []
    for payment in unapproved_payments:
        client = payment.client
        data.append({
            'Client Name': client.name,
            'contact1': client.phoneNumber1,
            'contact2': client.phoneNumber2,
            'Location': client.location,
        })

    df = pd.DataFrame(data)
    # print(df)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=unapproved_payments.xlsx'
    df.to_excel(response, index=False)

    return response


# ======== DOWNLOAD EXCEL CONTAINING EMPLOYEES WITH THE CLIENTS THEY BROUGHT INCLUDING THERE PAYMENT ========
def get_employee_with_clients_and_payments(request):
    employees = Employees.objects.all()

    data = []
    for employee in employees:
        clients = Client.objects.filter(employee=employee)
        if not clients.exists():
            data.append({
                'employee_name': employee.user.get_full_name(),
                'user_name': employee.user.username,
                'user_email': employee.user.email,
                'employee_contact': employee.phone,
                'client_name': None,
                'client_contact1': None,
                'client_contact2': None,
                'client_location': None,
                'amount_paid': None,
                'expected_amount': None,
                'remaining_amount': None,
                'installment_number': None,
                'total_Installments': None,
                'payment_date': None,
                'payment_approved': None,
            })
        else:
            for client in clients:
                payments = Payment.objects.filter(client=client)
                if not payments.exists():
                    data.append({
                        'employee_name': employee.user.get_full_name(),
                        'user_name': employee.user.username,
                        'user_email': employee.user.email,
                        'employee_contact': employee.phone,
                        'client_name': client.name,
                        'client_contact1': client.phoneNumber1,
                        'client_contact2': client.phoneNumber2,
                        'client_location': client.location,
                        'amount_paid': None,
                        'expected_amount': None,
                        'remaining_amount': None,
                        'installment_number': None,
                        'total_Installments': None,
                        'payment_date': None,
                        'payment_approved': None,
                    })
                else:
                    for payment in payments:
                        data.append({
                            'employee_name': employee.user.get_full_name(),
                            'user_name': employee.user.username,
                            'user_email': employee.user.email,
                            'employee_contact': employee.phone,
                            'client_name': client.name,
                            'client_contact1': client.phoneNumber1,
                            'client_contact2': client.phoneNumber2,
                            'client_location': client.location,
                            'amount_paid': payment.amount_paid,
                            'expected_amount': payment.total_amount,
                            'remaining_amount': payment.remaining_amount,
                            'installment_number': payment.installment_number,
                            'total_Installments': payment.total_installments,
                            'payment_date': payment.timestamp,
                            'payment_approved': payment.approved,
                        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=all_employees_clients_and_payments.xlsx'
    df.to_excel(response, index=False)

    return response


#  ============ UPLOADING EXCEL FILE ================
def upload_file(request):
    print("9005", "\n", "-"*30)
    upload_file_ = request.FILES['drive_file']
    print(upload_file_)
    ret = {}
    if upload_file_:
        filename = request.POST['filename']
        print(filename)
        sheet_name_ = request.POST['sheet_name']
        print(sheet_name_)
        # print("9005-1", filename, "\n", "-"*30)
        app_ = request.POST['app']
        # print("9005-2", app_, "\n", "-"*30)
        obj_name_ = request.POST['obj_name']
        function_name_ = request.POST['function_name']
        topic_id_ = request.POST['topic_id']
        folder_type_ = request.POST['folder_type']
        entity_name_ = request.POST['entity_name']
        if folder_type_ == "media":
            print("media")
            data_dir = settings.MEDIA_ROOT + '/' + app_ + '/' + topic_id_
            upload_file_ = request.FILES['drive_file']
            os.makedirs(data_dir, exist_ok=True)
            filename = request.POST['filename']
            file_path = os.path.join(data_dir, filename)

            model_name = request.POST['model_name']
            if model_name !="":
                record_id = request.POST['record_id']
                model_field_name = request.POST['model_field_name']
                model = apps.get_model(app_label=app_, model_name=model_name)
                obj = model.objects.get(id=record_id)
                s = "obj."+model_field_name+"='"+filename+"'"
                exec(s)
                obj.save()
                # print(file_path)
            with open(file_path, 'wb+') as destination:
                for c in upload_file_.chunks():
                    destination.write(c)
            ret['status'] = "ok"

            return HttpResponse(json.dumps(ret))
        #
        try:
            cube_dic = {}
            dimensions_ = request.POST['dimensions']
            fields_ = request.POST['fields']
            fact_model_field_ = request.POST['fact_model_field']
            # print("9005-161", fact_model_field_, "\n", "-"*30)
            #
            dimensions_s = dimensions_.split(",")
            # time_dim,country_dim,measure_dim
            # year,country_name,measure_name
            # WorldBankFact,amount
            fields_s = fields_.split(",")
            # print(fields_)
            # print(fact_model_field_)
            fact_model_field_s = fact_model_field_.split(",")
            cube_dic = {"dimensions": {}, "fact": {"model": fact_model_field_s[0], "field_name": fact_model_field_s[1]}}
            for j in range(len(dimensions_s)):
                f = fields_s[j]
                d = dimensions_s[j]
                dm = d.replace("_", "")
                cube_dic["dimensions"][d] = {"model": dm, "field_name": f}

            # print("90876 upload file: ")
            # print(cube_dic)
            # cube_dic = {"dimensions": {"time_dim": {"model": "TimeDim", "field_name": "year"},
            #                            "country_dim": {"model": "CountryDim", "field_name": "country_name"},
            #                            "measure_dim": {"model": "MeasureDim", "field_name": "measure_name"}},
            #             "fact": {"model": "WorldBankFact", "field_name": "amount"}
            #             }
            # print("90876-1 upload file: ")
            # print(cube_dic)
        except Exception as ex:
            pass

        add_dic = {"obj": obj_name_, "app": app_, "fun": function_name_,
                   "params": {"request": request, "folder_type": folder_type_, "sheet_name": sheet_name_,
                              "app": app_, "cube_dic": cube_dic},
                   "obj_param": {"topic_id": topic_id_, "app": app_, "entity_name": entity_name_}}

        try:
            add_dic["obj_param"]["entity_model"] = cube_dic["dimensions"][entity_name_+"_dim"]["model"]
        except Exception as ex:
            pass
        try:
            pass
            add_dic["obj_param"]["measure_model"] = cube_dic["dimensions"]["measure_dim"]["model"]
        except Exception as ex:
            pass

        # print("9010")
        # print(add_dic)
        # print("9010")
        return activate_obj_function(request, add_dic)
        # print(ret
        # print("9011")
        # ret['status'] = "ok"
    else:
        ret['status'] = "ko"

    # print("ret", ret)
    return HttpResponse(json.dumps(ret))


#  ============ ASSISSTING FUNCTION TO UPLOAD EXCEL FILE ================
def get_data_link(request):
    errors = ""
    dic_ = request.POST["dic"]
    try:
        pass
        print('\n 9050-150-50 core views get_data_link dic_ ', "\n", dic_,"\n", "-"*100)
    except Exception as ex:
        print(str(ex))

    try:
        pass
        print("get_data_link 99999: "+dic_)
        dic_ = eval(dic_)
    except Exception as ex:
        print("error 4562-22-1",str(ex))

    # errors += "1 "
    # log_debug("get_data_link 99999: "+errors)
    parent_model_fk_name=""
    try:
        parent_model_fk_name=dic_["parent_model_fk_name"]
    except Exception as ex:
        pass

    parent_model_fk_name=""
    try:
        parent_model_fk_name=dic_["parent_model_fk_name"]
    except Exception as ex:
        pass

    # print("parent_model_fk_name\n", parent_model_fk_name, "\nparent_model_fk_name")

    multiple_select_fields = None
    if "multiple_select_fields" in dic_:
        if len(dic_["multiple_select_fields"]) > 0:
            multiple_select_fields = dic_["multiple_select_fields"]
    app_ = dic_['app']
    model_ = dic_['model']
    # print("model_: "+model_)
    if model_ == "":
        dic = {'status': 'ko', "dic": {}}
        return JsonResponse(dic)

    model = apps.get_model(app_label=app_, model_name=model_)
    # for testing only --
    # model__ = apps.get_model(app_label=app_, model_name="XBRLDimCompany")
    # df = pd.DataFrame(model__.objects.all().values())
    # print(df)
    # for index, row in df.iterrows():
    #     print(index)
    #     # print(row, "\n", row["full_name"])
    #     print(row["sic_code"])
    # for testing only --
    p_key_field_name = model._meta.pk.name
    print("p_key_field_name===== ", p_key_field_name)
    print("dic_[fields]====== ", dic_["fields"])
    if p_key_field_name not in dic_["fields"]:
        dic_["fields"].insert(0, p_key_field_name)
    fields_str = '"'
    for f in dic_["fields"]:
        try:
            if f != "":
                exec(f + ' = []')
                fields_str += f + '","'
        except Exception as ex:
            print("error 4000-1: "+str(ex))
    fields_str = fields_str[:len(fields_str)-2]
    # print("9030","\n", fields_str,"\n","=2"*50)
    # fields_ = model._meta.get_fields(include_parents=True, include_hidden=True)
    # print(fields_)
    number_of_rows_ = 2
    try:
        number_of_rows_ = dic_['number_of_rows']
        number_of_rows_ = int(number_of_rows_)
    except Exception as ex:
        pass
        # print(ex)
    parent_id_ = -1
    try:
        parent_id_ = int(dic_['parent_id'])
    except Exception as ex:
        # print("error 500 "+str(ex))
        pass
    try:
        company_obj_id_ = dic_['company_obj_id']
    except Exception as ex:
        print("error 440: "+str(ex))
    filters = dic_['filters']
    if len(dic_['order_by']) > 0:
        order_by = dic_['order_by']
    else:
        order_by = ""

    # print(" company_obj_id_", company_obj_id_)

    if company_obj_id_ != "-1" and company_obj_id_ != -1:
        # log_debug("get_data_link company_obj_id_: "+str(company_obj_id_))
        s = 'model.objects'
        s_ = ''
        try:
            parent_model = apps.get_model(app_label=app_, model_name=app_+"web")
            print(parent_model)
            company_obj = parent_model.objects.get(id=company_obj_id_)
            if model.model_field_exists(app_+'_web') and isinstance(model._meta.get_field(app_+'_web'),
                                                                    ForeignKey):
                s_ += app_ + '_web=company_obj '
            if parent_id_ > -1:
                parent_model_ = dic_['parent_model']
                parent_pkey_ = parent_id_
                parent_model__ = apps.get_model(app_label=app_, model_name=parent_model_)

                ss_=parent_model_
                if ss_[len(ss_)-1] == "s":
                    ss_ = parent_model_[:-1]
                if parent_model_fk_name == "":
                    parent_model_fk_name = ss_
                pk = parent_model__._meta.pk.name
                parent_obj__ = eval('parent_model__.objects.get('+pk+'=parent_pkey_)')
                if s_ != '':
                    s_ += ', '
                s_ += parent_model_fk_name+'=parent_obj__'
                # print("sss1", "\n", s, "\n", sss1")
        except Exception as ex:
            print(ex)
        if s_ != '':
            s += '.filter('+s_+')'
        # print('s00111\n', s, '\ns00111')
    else:
        if parent_id_ > -1:
            parent_model_ = dic_['parent_model']
            parent_pkey_ = parent_id_
            parent_model__ = apps.get_model(app_label=app_, model_name=parent_model_)

            if parent_model_fk_name == "":
                parent_model_fk_name = parent_model_[:-1]

            # parent_obj__ = parent_model__.objects.get(id=parent_pkey_)

            parent_obj__ = eval('parent_model__.objects.get('+parent_model__._meta.pk.name+'=parent_pkey_)')

            # print("parent_obj__\n", parent_obj__)

            s = 'model.objects.filter(' + parent_model_fk_name+'=parent_obj__)'
        else:
            s = 'model.objects'
        # print('90500 s '+s)
    # print("9030-2\n", s)
    try:
        for f in filters:
            filter_field_ = f  # filters[f]["filter_field"] #
            filter_value_ = str(filters[f]["value"])
            filter_field_a = ""
            try:
                filter_field_a = str(filters[f]["field"])
            except Exception as exx:
                pass
            foreign_table_ = ""
            try:
                foreign_table_ = filters[f]["foreign_table"]
            except Exception as exx:
                pass
            if filter_value_ != "":
                # print(foreign_table_)
                if foreign_table_ != "":
                    # print(1111111111)
                    if filter_field_a != "":
                        # need need need to check this one. I changed it and it might have effect on other reports
                        filter_field_ = filter_field_a
                    # print(filter_field_)
                    # print(111122222)
                    # f__ = model._meta.get_field(filter_field_)
                    # print(f__)
                    # t__ = f__.get_internal_type()
                    # print(t__)
                    # print(str(t__))
                    # s += '.filter('+foreign_table_+'__'+filter_field_+'='+filter_value_+')'
                    # if str(t__)=="AutoField":
                    #     print(3333333)
                    index = filter_field_.find("id")
                    if index != -1:
                        s += '.filter(' + foreign_table_ + '__' + filter_field_ + '=' + filter_value_ + ')'
                    else:
                        # print(44444555)
                        s += '.filter(' + foreign_table_ + '__' + filter_field_ + '__icontains="'+filter_value_+'")'
                else:
                    # print(22222222222)
                    if filter_field_ == "id":
                        #s += '.filter('+filter_field_+'__icontains='+filter_value_+')'
                        s += '.filter('+filter_field_+'='+filter_value_+')'
                    else:
                        s += '.filter('+filter_field_+'__icontains="'+filter_value_+'")'
        # print(s)
        # print("9030-22")
        n_ = -1
        try:
            primary_key_list_filter_ = dic_["primary_key_list_filter"]
            n_ = len(primary_key_list_filter_)
            if primary_key_list_filter_ and n_ > 0:
                s += '.filter('+p_key_field_name+'__in=primary_key_list_filter_)'
        except Exception as ex:
            ("Error 90855-23 "+str(ex))

        # print("9030-221")

        if order_by != "":
            s += '.order_by("'+order_by["field"]+'")'
            if order_by["direction"] == "descending":
                s += '.reverse()'
        if multiple_select_fields:
            ss__ = s+'.all()[:number_of_rows_]'
            # print('ss__ for data__')
            # print(ss__)
            # print('ss__')
            data__ = eval(ss__)

        # print("9030-222")
        s += '.all()[:number_of_rows_].values('+fields_str+')'

        print("="*100, '\ns111-1 for d_data\n', "\ns=", number_of_rows_, s, "=\n", "="*100, "\n")

        d_data = eval(s)
        # print("d_data\n", d_data)
    except Exception as ex:
        print("3030-1 core error 300 "+str(ex))
        # pass
    dic = {}
    if multiple_select_fields:
        for z in multiple_select_fields:
            print("=====OPIO======  ", z)
            dic[z] = []
            for q in data__:
                model_z_name = z+"s"
                model_z = apps.get_model(app_label=app_, model_name=model_z_name)
                df = pd.DataFrame(model_z.objects.all().values())
                p_key_field_z_name = model_z._meta.pk.name
                if p_key_field_z_name != "id":
                    p_key_field_z_name = p_key_field_z_name + "_id"
                qs = eval('q.'+z+'.all()')
                s = ""
                for q_ in qs:
                    if s != "":
                        s += ","
                    s += str(eval('q_.'+p_key_field_z_name))
                dic[z].append(s)
    try:
        for q in d_data:
            for f in dic_["fields"]:
                # print("fffff====== ", f)
                if f != "":
                    # print(f+'.append(q[\''+f+'\'])')
                    kk__ = q[f]
                    if (not (isinstance(q[f], float) or isinstance(q[f], int))) and isinstance(kk__, numbers.Number):
                        # print(isinstance(q[f], numbers.Number))
                        kk__ = float(kk__)
                    # eval(f+'.append(q[\''+f+'\'])')
                    eval(f+'.append(kk__)')
                    # print(eval(f))
        for ff in dic_["fields"]:
            if ff != "":
                dic[ff] = eval(ff)
    except Exception as ex:
        pass
    dic["pkf_name"] = p_key_field_name
    # print(dic, "=2"*50)
    dic = {'status': 'ok', "dic": dic}
    # print('core view 9055 get_data_link dic_= ', dic)
    return JsonResponse(dic)